# services/relatorios_service.py
import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# Lista de atributos FM26 (copiada do sistema original)
ATRIBUTOS_FM26 = [
    'escanteios', 'cruzamentos', 'drible', 'finalizacao', 'primeiro_controle',
    'cobranca_faltas', 'cabecada', 'chutes_longe', 'arremessos_laterais',
    'marcacao', 'passe', 'cobranca_penaltis', 'desarme', 'tecnica',
    'agressividade', 'antecipacao', 'coragem', 'composicao', 'concentracao',
    'decisao', 'determinacao', 'criatividade', 'lideranca', 'movimentacao_sem_bola',
    'posicionamento', 'trabalho_equipe', 'visao_jogo', 'intensidade_trabalho',
    'aceleracao', 'agilidade', 'equilibrio', 'altura_salto', 'condicao_fisica_natural',
    'velocidade_maxima', 'resistencia', 'forca_fisica', 'reflexos', 'jogo_aereo_goleiro',
    'defesas_goleiro', 'comando_area', 'comunicacao_goleiro', 'chutes_goleiro',
    'um_contra_um_goleiro', 'saida_gol', 'tendencia_socar', 'arremessos_goleiro',
    'excentricidade', 'consistencia', 'jogo_sujo', 'jogos_importantes',
    'propensao_lesao', 'versatilidade', 'adaptabilidade', 'ambicao', 'lealdade',
    'pressao', 'profissionalismo', 'esportividade', 'temperamento', 'controversia'
]

# =========================================================================
# RELATÓRIO PARA DIRETORIA (Executivo)
# =========================================================================
def gerar_relatorio_diretoria(df_jogadores, df_comissao):
    """
    Gera um relatório executivo em Excel para a diretoria.
    """
    output = io.BytesIO()
    wb = Workbook()

    # Cria a primeira aba explicitamente (evita None)
    ws = wb.create_sheet("Resumo Geral")
    ws['A1'] = "RELATÓRIO EXECUTIVO - VILAVELHENSE FC"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    row = 3
    ws[f'A{row}'] = "Indicador"
    ws[f'B{row}'] = "Valor"
    row += 1

    ws[f'A{row}'] = "Total de Jogadores"
    ws[f'B{row}'] = len(df_jogadores)
    row += 1
    ws[f'A{row}'] = "Total de Membros da Comissão"
    ws[f'B{row}'] = len(df_comissao)
    row += 1

    if 'Idade' in df_jogadores.columns:
        ws[f'A{row}'] = "Idade Média dos Jogadores"
        ws[f'B{row}'] = round(df_jogadores['Idade'].mean(), 1)
        row += 1
    if 'IMC' in df_jogadores.columns:
        ws[f'A{row}'] = "IMC Médio"
        ws[f'B{row}'] = round(df_jogadores['IMC'].mean(), 1)
        row += 1
    if 'Gordura_Corporal_%' in df_jogadores.columns:
        ws[f'A{row}'] = "% Gordura Médio"
        ws[f'B{row}'] = round(df_jogadores['Gordura_Corporal_%'].mean(), 1)
        row += 1
    if 'Rating_Geral_FM26' in df_jogadores.columns:
        ws[f'A{row}'] = "Rating Geral Médio (FM26)"
        ws[f'B{row}'] = round(df_jogadores['Rating_Geral_FM26'].mean(), 1)
        row += 1

    # Distribuição por posição
    row += 2
    ws[f'A{row}'] = "Distribuição por Posição"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    if 'Posicao_Principal' in df_jogadores.columns:
        for pos, qtd in df_jogadores['Posicao_Principal'].value_counts().items():
            ws[f'A{row}'] = pos
            ws[f'B{row}'] = qtd
            row += 1

    # ===== PÁGINA 2: Análise Física =====
    ws2 = wb.create_sheet("Análise Física")
    ws2['A1'] = "ANÁLISE FÍSICA DOS JOGADORES"
    ws2['A1'].font = Font(size=14, bold=True)
    ws2.merge_cells('A1:D1')

    row = 3
    ws2[f'A{row}'] = "Jogador"
    ws2[f'B{row}'] = "IMC"
    ws2[f'C{row}'] = "% Gordura"
    ws2[f'D{row}'] = "Estado Físico"
    for col in 'ABCD':
        ws2[col+str(row)].font = Font(bold=True)
    row += 1

    for _, jog in df_jogadores.iterrows():
        ws2[f'A{row}'] = jog.get('nome_completo', 'N/I')
        ws2[f'B{row}'] = jog.get('IMC', 'N/I') if pd.notna(jog.get('IMC')) else 'N/I'
        ws2[f'C{row}'] = jog.get('Gordura_Corporal_%', 'N/I') if pd.notna(jog.get('Gordura_Corporal_%')) else 'N/I'
        ws2[f'D{row}'] = jog.get('Estado_Fisico', 'N/I')
        row += 1

    # ===== PÁGINA 3: Destaques =====
    ws3 = wb.create_sheet("Destaques")
    ws3['A1'] = "JOGADORES EM DESTAQUE"
    ws3['A1'].font = Font(size=14, bold=True)
    ws3.merge_cells('A1:D1')

    row = 3
    if 'Rating_Geral_FM26' in df_jogadores.columns and not df_jogadores.empty:
        melhor = df_jogadores.loc[df_jogadores['Rating_Geral_FM26'].idxmax()]
        ws3[f'A{row}'] = "Melhor Rating Geral"
        ws3[f'B{row}'] = f"{melhor.get('nome_completo', 'N/I')} - {melhor['Rating_Geral_FM26']:.1f}"
        row += 1

    if 'Idade' in df_jogadores.columns and not df_jogadores.empty:
        mais_jovem = df_jogadores.loc[df_jogadores['Idade'].idxmin()]
        ws3[f'A{row}'] = "Jogador mais jovem"
        ws3[f'B{row}'] = f"{mais_jovem.get('nome_completo', 'N/I')} - {mais_jovem['Idade']} anos"
        row += 1

    if 'Estado_Fisico' in df_jogadores.columns:
        otimos = df_jogadores[df_jogadores['Estado_Fisico'] == 'otimo']
        if not otimos.empty:
            ws3[f'A{row}'] = "Jogadores com estado físico 'Ótimo'"
            row += 1
            for _, jog in otimos.iterrows():
                ws3[f'A{row}'] = jog.get('nome_completo', 'N/I')
                row += 1

    # ===== Formatação geral =====
    for ws_temp in [ws, ws2, ws3]:
        for col in ws_temp.columns:
            max_length = 0
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            if col:
                col_letter = col[0].column_letter
                ws_temp.column_dimensions[col_letter].width = adjusted_width

    wb.save(output)
    output.seek(0)
    return output


# =========================================================================
# RELATÓRIO INDIVIDUAL DO JOGADOR
# =========================================================================
def gerar_relatorio_jogador(jogador_row):
    """
    Gera relatório individual de um jogador em Excel.
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.create_sheet("Relatório Jogador")

    ws['A1'] = f"RELATÓRIO INDIVIDUAL - {jogador_row.get('nome_completo', 'Jogador')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    row = 3
    campos_pessoais = [
        ('Nome Completo', 'nome_completo'),
        ('Apelido', 'apelido'),
        ('Data Nascimento', 'data_nascimento'),
        ('Idade', 'Idade'),
        ('Posição Principal', 'Posicao_Principal'),
        ('Pé Preferido', 'pe_pref'),
        ('Altura (cm)', 'altura_cm'),
        ('Peso (kg)', 'peso_kg'),
        ('IMC', 'IMC'),
        ('% Gordura', 'Gordura_Corporal_%'),
        ('Massa Magra (kg)', 'Massa_Magra_kg'),
        ('Estado Físico', 'Estado_Fisico'),
        ('Lesionado', 'lesionado'),
    ]

    for label, key in campos_pessoais:
        valor = jogador_row.get(key, 'N/I')
        if pd.isna(valor) or valor == '':
            valor = 'N/I'
        ws[f'A{row}'] = label
        ws[f'B{row}'] = valor
        row += 1

    # Atributos FM26 (top 15)
    row += 2
    ws[f'A{row}'] = "PRINCIPAIS ATRIBUTOS FM26"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    atributos_existentes = []
    for attr in ATRIBUTOS_FM26:
        val = jogador_row.get(attr, np.nan)
        if pd.notna(val):
            atributos_existentes.append((attr, val))

    atributos_existentes.sort(key=lambda x: x[1], reverse=True)
    for attr, val in atributos_existentes[:15]:
        ws[f'A{row}'] = attr.replace('_', ' ').title()
        ws[f'B{row}'] = val
        row += 1

    wb.save(output)
    output.seek(0)
    return output


# =========================================================================
# RELATÓRIO DA COMISSÃO TÉCNICA (individual)
# =========================================================================
def gerar_relatorio_comissao(membro_row):
    """
    Gera relatório individual de um membro da comissão.
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.create_sheet("Relatório Comissão")

    ws['A1'] = f"RELATÓRIO - {membro_row.get('nome', 'Membro da Comissão')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    row = 3
    campos = [
        ('Nome', 'nome'),
        ('Cargo', 'cargo'),
        ('Idade', 'idade'),
        ('Data Nascimento', 'data_nascimento'),
        ('Cidade/UF', 'cidade_uf'),
        ('País', 'pais'),
    ]
    for label, key in campos:
        valor = membro_row.get(key, 'N/I')
        if pd.isna(valor) or valor == '':
            valor = 'N/I'
        ws[f'A{row}'] = label
        ws[f'B{row}'] = valor
        row += 1

    row += 2
    ws[f'A{row}'] = "HISTÓRICO PROFISSIONAL"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    historico = membro_row.get('historico_profissional', '')
    ws[f'A{row}'] = historico if historico else "Nenhum histórico registrado."

    # Atributos de staff (se houver)
    if 'atributos_staff' in membro_row or any(col.startswith('staff_') for col in membro_row.index):
        row += 2
        ws[f'A{row}'] = "ATRIBUTOS DE STAFF"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for col in membro_row.index:
            if col.startswith('staff_') or col in ['lideranca', 'motivacao', 'tatica', 'treino']:
                val = membro_row.get(col, 'N/I')
                if pd.notna(val) and val != '':
                    ws[f'A{row}'] = col.replace('_', ' ').title()
                    ws[f'B{row}'] = val
                    row += 1

    wb.save(output)
    output.seek(0)
    return output


# =========================================================================
# RELATÓRIO COMPLETO DA COMISSÃO (com análise tática e escalação)
# =========================================================================
def gerar_relatorio_comissao_completo(membro_row, df_jogadores, formacao, estilo):
    """
    Gera relatório completo com análise do elenco, melhores jogadores e escalação sugerida.
    """
    output = io.BytesIO()
    wb = Workbook()

    # ===== ABA 1: Dados do Membro =====
    ws_membro = wb.create_sheet("Membro")
    ws_membro['A1'] = f"RELATÓRIO - {membro_row.get('nome', 'Membro da Comissão')}"
    ws_membro['A1'].font = Font(size=16, bold=True)
    ws_membro.merge_cells('A1:D1')

    row = 3
    campos = [
        ('Nome', 'nome'),
        ('Cargo', 'cargo'),
        ('Idade', 'idade'),
        ('Data Nascimento', 'data_nascimento'),
        ('Cidade/UF', 'cidade_uf'),
        ('País', 'pais'),
    ]
    for label, key in campos:
        valor = membro_row.get(key, 'N/I')
        if pd.isna(valor) or valor == '':
            valor = 'N/I'
        ws_membro[f'A{row}'] = label
        ws_membro[f'B{row}'] = valor
        row += 1

    # ===== ABA 2: Análise do Elenco =====
    ws_elenco = wb.create_sheet("Análise do Elenco")
    ws_elenco['A1'] = "ANÁLISE GERAL DO ELENCO"
    ws_elenco['A1'].font = Font(size=14, bold=True)
    ws_elenco.merge_cells('A1:D1')

    row = 3
    ws_elenco[f'A{row}'] = "Total de Jogadores"
    ws_elenco[f'B{row}'] = len(df_jogadores)
    row += 1
    if 'Idade' in df_jogadores.columns:
        ws_elenco[f'A{row}'] = "Idade Média"
        ws_elenco[f'B{row}'] = round(df_jogadores['Idade'].mean(), 1)
        row += 1
    if 'IMC' in df_jogadores.columns:
        ws_elenco[f'A{row}'] = "IMC Médio"
        ws_elenco[f'B{row}'] = round(df_jogadores['IMC'].mean(), 1)
        row += 1
    if 'Rating_Geral_FM26' in df_jogadores.columns:
        ws_elenco[f'A{row}'] = "Rating Geral Médio"
        ws_elenco[f'B{row}'] = round(df_jogadores['Rating_Geral_FM26'].mean(), 1)
        row += 1

    # Distribuição por posição
    row += 2
    ws_elenco[f'A{row}'] = "Distribuição por Posição"
    ws_elenco[f'A{row}'].font = Font(bold=True)
    row += 1
    if 'Posicao_Principal' in df_jogadores.columns:
        for pos, qtd in df_jogadores['Posicao_Principal'].value_counts().items():
            ws_elenco[f'A{row}'] = pos
            ws_elenco[f'B{row}'] = qtd
            row += 1

    # ===== ABA 3: Melhores por Posição =====
    ws_melhores = wb.create_sheet("Melhores por Posição")
    ws_melhores['A1'] = "MELHORES JOGADORES POR POSIÇÃO"
    ws_melhores['A1'].font = Font(size=14, bold=True)
    ws_melhores.merge_cells('A1:D1')

    row = 3
    posicoes = ['Goleiro', 'Zagueiro', 'Lateral', 'Volante', 'Meia-Central', 'Meia-Atacante', 'Ponta', 'Centroavante']
    for pos in posicoes:
        df_pos = df_jogadores[df_jogadores['Posicao_Principal'] == pos]
        if not df_pos.empty:
            df_pos = df_pos.sort_values('Rating_Geral_FM26', ascending=False)
            ws_melhores[f'A{row}'] = pos.upper()
            ws_melhores[f'A{row}'].font = Font(bold=True)
            row += 1
            for _, jog in df_pos.head(5).iterrows():
                ws_melhores[f'A{row}'] = jog.get('nome_completo', 'N/I')
                ws_melhores[f'B{row}'] = jog.get('Rating_Geral_FM26', 0)
                row += 1
            row += 1

    # ===== ABA 4: Escalação Sugerida (baseada em rating) =====
    ws_escalacao = wb.create_sheet("Escalação Sugerida")
    ws_escalacao['A1'] = f"ESCALAÇÃO SUGERIDA - {formacao} | ESTILO: {estilo}"
    ws_escalacao['A1'].font = Font(size=14, bold=True)
    ws_escalacao.merge_cells('A1:E1')

    # Escalação simples: pega os 11 melhores ratings, depois os próximos 10 como reservas
    df_jogadores_sorted = df_jogadores.sort_values('Rating_Geral_FM26', ascending=False)
    titulares = []
    reservas = []

    for _, row in df_jogadores_sorted.head(11).iterrows():
        titulares.append({
            'nome': row.get('nome_completo', 'N/I'),
            'apelido': row.get('apelido', ''),
            'posicao': row.get('Posicao_Principal', ''),
            'rating': row.get('Rating_Geral_FM26', 0)
        })
    for _, row in df_jogadores_sorted.iloc[11:21].iterrows():
        reservas.append({
            'nome': row.get('nome_completo', 'N/I'),
            'apelido': row.get('apelido', ''),
            'posicao': row.get('Posicao_Principal', ''),
            'rating': row.get('Rating_Geral_FM26', 0)
        })

    row = 3
    ws_escalacao[f'A{row}'] = "TITULARES"
    ws_escalacao[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_escalacao[f'A{row}'] = "Jogador"
    ws_escalacao[f'B{row}'] = "Posição"
    ws_escalacao[f'C{row}'] = "Rating"
    for col in 'ABC':
        ws_escalacao[col+str(row)].font = Font(bold=True)
    row += 1
    for j in titulares:
        ws_escalacao[f'A{row}'] = j['nome']
        ws_escalacao[f'B{row}'] = j['posicao']
        ws_escalacao[f'C{row}'] = j['rating']
        row += 1

    row += 2
    ws_escalacao[f'A{row}'] = "RESERVAS"
    ws_escalacao[f'A{row}'].font = Font(bold=True)
    row += 1
    for j in reservas:
        ws_escalacao[f'A{row}'] = j['nome']
        ws_escalacao[f'B{row}'] = j['posicao']
        ws_escalacao[f'C{row}'] = j['rating']
        row += 1

    # ===== Formatação das abas =====
    for ws_temp in [ws_membro, ws_elenco, ws_melhores, ws_escalacao]:
        for col in ws_temp.columns:
            max_length = 0
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            if col:
                col_letter = col[0].column_letter
                ws_temp.column_dimensions[col_letter].width = adjusted_width

    wb.save(output)
    output.seek(0)
    return output